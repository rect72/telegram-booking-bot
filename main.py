import asyncio
import logging
import os
import re
from datetime import datetime
from pathlib import Path

from aiohttp import ClientTimeout, TCPConnector
from aiogram import Bot, Dispatcher
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
    BotCommand,
    FSInputFile,
    ReplyKeyboardRemove,  # ✅ добавили
)
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from config import BOT_TOKEN

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
)

logging.getLogger("aiohttp").setLevel(logging.INFO)
logging.getLogger("aiogram").setLevel(logging.INFO)

dp = Dispatcher()

ADMIN_CHAT_ID = os.getenv("ADMIN_CHAT_ID", "").strip()

services_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Стрижка")],
        [KeyboardButton(text="Консультация")],
        [KeyboardButton(text="Диагностика")],
        [KeyboardButton(text="Окрашивание")],
        [KeyboardButton(text="Запись на звонок")],
    ],
    resize_keyboard=True,
)

cancel_keyboard = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text="Отменить")]],
    resize_keyboard=True,
)

time_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="10:00")],
        [KeyboardButton(text="12:00")],
        [KeyboardButton(text="14:00")],
        [KeyboardButton(text="16:00")],
    ],
    resize_keyboard=True,
)

confirm_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Подтвердить")],
        [KeyboardButton(text="Изменить")],
        [KeyboardButton(text="Отменить")],
    ],
    resize_keyboard=True,
)

ADMIN_KEYBOARD = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Количество заявок")],
        [KeyboardButton(text="Последние заявки")],
        [KeyboardButton(text="Выгрузить Excel")],
    ],
    resize_keyboard=True,
)

SERVICES = ["Стрижка", "Консультация", "Диагностика", "Окрашивание", "Запись на звонок"]

BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = str(BASE_DIR / "заявки.xlsx")
EXCEL_SHEET = "Заявки"
EXCEL_HEADERS = ["timestamp", "service", "name", "phone", "datetime_text", "comment"]


def is_valid_phone(phone: str) -> bool:
    """
    Разрешаем ТОЛЬКО:
    +7XXXXXXXXXX (11 цифр после +7) или 8XXXXXXXXXX (11 цифр, первая 8)
    Т.е. строго 12 символов с + (пример +79991234567) или 11 символов без + (89991234567)
    """
    if not phone:
        return False

    s = phone.strip().replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    return bool(re.fullmatch(r"(\+7\d{10}|8\d{10})", s))


def is_valid_datetime_text(text: str) -> bool:
    if not text or not text.strip():
        return False
    t = text.strip()
    return bool(
        re.search(r"\d{1,2}\s*[^\d]*\d{1,2}:\d{2}", t, re.IGNORECASE)
        or re.search(r"\d{1,2}:\d{2}", t)
    )


def is_valid_date_text(text: str) -> bool:
    if not text or not text.strip():
        return False

    t = text.strip().lower()

    months = (
        "января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря|"
        "янв|фев|мар|апр|май|июн|июл|авг|сен|окт|ноя|дек"
    )

    pattern = rf"^\s*(0?[1-9]|[12]\d|3[01])\s+({months})(\s+\d{{4}})?\s*$"
    return bool(re.match(pattern, t, re.IGNORECASE))


async def notify_admin(bot: Bot, text: str) -> None:
    if not ADMIN_CHAT_ID:
        return
    try:
        await bot.send_message(chat_id=int(ADMIN_CHAT_ID), text=text)
    except Exception:
        logging.exception("Не удалось отправить заявку админу (бот не падает)")


def has_empty_required_fields(data: dict) -> tuple[bool, str]:
    required = ["service", "name", "phone", "datetime", "comment"]
    for key in required:
        v = data.get(key)
        if v is None:
            return True, f"Поле '{key}' пустое."
        if isinstance(v, str) and not v.strip():
            return True, f"Поле '{key}' пустое."
    return False, ""


def save_order_to_excel(data: dict) -> None:
    try:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        try:
            wb = load_workbook(EXCEL_PATH)
        except FileNotFoundError:
            wb = Workbook()
        except InvalidFileException:
            wb = Workbook()
        except PermissionError as e:
            logging.error(f"Excel недоступен (PermissionError): {e}")
            return
        except OSError as e:
            logging.error(f"Excel недоступен (OSError): {e}")
            return

        if EXCEL_SHEET in wb.sheetnames:
            ws = wb[EXCEL_SHEET]
        else:
            ws = wb.create_sheet(EXCEL_SHEET)
            ws.append(EXCEL_HEADERS)

        if ws.max_row == 0 or (ws.max_row == 1 and ws.cell(1, 1).value is None):
            ws.append(EXCEL_HEADERS)

        ws.append(
            [
                ts,
                data.get("service"),
                data.get("name"),
                data.get("phone"),
                data.get("datetime"),
                data.get("comment", "Нет"),
            ]
        )

        try:
            wb.save(EXCEL_PATH)
        except PermissionError as e:
            logging.error(f"Не удалось сохранить Excel (PermissionError): {e}")
            return
        except OSError as e:
            logging.error(f"Не удалось сохранить Excel (OSError): {e}")
            return

    except Exception:
        logging.exception("save_order_to_excel: неожиданная ошибка, заявка не сохранена")


def load_orders_from_excel(limit: int = 10) -> list[list[str]]:
    try:
        wb = load_workbook(EXCEL_PATH)
        if EXCEL_SHEET not in wb.sheetnames:
            return []

        ws = wb[EXCEL_SHEET]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        data_rows = rows[1:] if len(rows) > 1 else []
        data_rows = data_rows[-limit:] if limit else data_rows

        result: list[list[str]] = []
        for r in data_rows:
            result.append([("" if v is None else str(v)) for v in r])

        return result

    except FileNotFoundError:
        return []
    except InvalidFileException:
        return []
    except PermissionError as e:
        logging.error(f"Excel PermissionError: {e}")
        return []
    except OSError as e:
        logging.error(f"Excel OSError: {e}")
        return []
    except Exception:
        logging.exception("load_orders_from_excel: неожиданная ошибка")
        return []


def get_orders_count() -> int:
    try:
        wb = load_workbook(EXCEL_PATH)
        if EXCEL_SHEET not in wb.sheetnames:
            return 0
        ws = wb[EXCEL_SHEET]
        if ws.max_row <= 1:
            return 0
        return ws.max_row - 1
    except Exception:
        return 0


class OrderForm(StatesGroup):
    waiting_for_name = State()
    waiting_for_phone = State()
    waiting_for_date = State()
    waiting_for_time = State()
    waiting_for_comment = State()
    waiting_for_confirmation = State()


def is_admin(user_id: int) -> bool:
    if not ADMIN_CHAT_ID:
        return False
    return str(user_id) == str(ADMIN_CHAT_ID)


@dp.message(CommandStart())
async def start_handler(message: Message):
    logging.info("Got /start")
    await message.answer(
        "Здравствуйте! Я помогу оставить заявку на услугу. Выберите услугу:",
        reply_markup=services_keyboard,
    )


@dp.message(Command("cancel"))
async def cancel_command(message: Message, state: FSMContext):
    await state.clear()
    await message.answer(
        "Действие отменено. Если хотите начать заново, нажмите /start.",
        reply_markup=services_keyboard,
    )


@dp.message(lambda m: m.text == "Отменить")
async def cancel_button(message: Message, state: FSMContext):
    await state.clear()
    await message.answer(
        "Заявка отменена. Чтобы начать заново, нажмите /start.",
        reply_markup=services_keyboard,
    )


@dp.message(lambda message: message.text in SERVICES)
async def choose_service(message: Message, state: FSMContext):
    await state.update_data(service=message.text)
    await state.set_state(OrderForm.waiting_for_name)
    await message.answer("Введите ваше имя:", reply_markup=cancel_keyboard)


@dp.message(OrderForm.waiting_for_name)
async def process_name(message: Message, state: FSMContext):
    name = (message.text or "").strip()
    if not name:
        await message.answer("Имя не должно быть пустым. Введите ваше имя:")
        return

    await state.update_data(name=name)
    await state.set_state(OrderForm.waiting_for_phone)
    await message.answer(
        "Введите ваш телефон. Пример: +79991234567 или 89991234567",
        reply_markup=cancel_keyboard,
    )


@dp.message(OrderForm.waiting_for_phone)
async def process_phone(message: Message, state: FSMContext):
    phone = (message.text or "").strip()

    if not is_valid_phone(phone):
        await message.answer(
            "Телефон введён неверно.\n"
            "Формат: +79991234567 или 89991234567 (ровно 11 цифр после 7/в начале 8).",
            reply_markup=cancel_keyboard,
        )
        return

    await state.update_data(phone=phone)
    await state.set_state(OrderForm.waiting_for_date)
    await message.answer(
        "Введите дату записи.\nНапример: 5 июля или 05 июля 2026",
        reply_markup=cancel_keyboard,
    )


@dp.message(OrderForm.waiting_for_date)
async def process_date(message: Message, state: FSMContext):
    date_text = (message.text or "").strip()

    if not date_text:
        await message.answer(
            "Дата не должна быть пустой.\nНапример: 5 июля",
            reply_markup=cancel_keyboard,
        )
        return

    if not is_valid_date_text(date_text):
        await message.answer(
            "Дата введена неверно.\n"
            "Пример: 5 июля или 5 июля 2026",
            reply_markup=cancel_keyboard,
        )
        return

    await state.update_data(date=date_text)
    await state.set_state(OrderForm.waiting_for_time)

    await message.answer("Выберите время:", reply_markup=time_keyboard)


@dp.message(OrderForm.waiting_for_time)
async def process_time(message: Message, state: FSMContext):
    time_text = (message.text or "").strip()
    allowed_times = {"10:00", "12:00", "14:00", "16:00"}

    if time_text not in allowed_times:
        await message.answer("Выберите время кнопками:", reply_markup=time_keyboard)
        return

    data = await state.get_data()
    date_text = data.get("date", "").strip()

    if not date_text:
        await state.set_state(OrderForm.waiting_for_date)
        await message.answer("Введите дату заново:", reply_markup=cancel_keyboard)
        return

    datetime_text = f"{date_text} {time_text}"

    if not is_valid_datetime_text(datetime_text):
        await message.answer(
            "Дата/время введены некорректно. Попробуйте ещё раз.\nВведите дату:",
            reply_markup=cancel_keyboard,
        )
        await state.set_state(OrderForm.waiting_for_date)
        return

    await state.update_data(datetime=datetime_text)
    await state.set_state(OrderForm.waiting_for_comment)

    await message.answer(
        "Введите комментарий к заявке.\n"
        "Если комментария нет, напишите: Нет",
        reply_markup=cancel_keyboard,
    )


@dp.message(OrderForm.waiting_for_comment)
async def process_comment(message: Message, state: FSMContext):
    comment = (message.text or "").strip()

    if not comment:
        comment = "Нет"
    elif comment.lower() in ("нет", "нет.", "no"):
        comment = "Нет"

    await state.update_data(comment=comment)

    data = await state.get_data()
    is_empty, err_text = has_empty_required_fields(data)
    if is_empty:
        await state.set_state(OrderForm.waiting_for_name)
        await message.answer(
            f"Не все обязательные поля заполнены: {err_text}\nВведите ваше имя:",
            reply_markup=cancel_keyboard,
        )
        return

    text = (
        "Проверьте заявку:\n"
        f"Услуга: {data.get('service')}\n"
        f"Имя: {data.get('name')}\n"
        f"Телефон: {data.get('phone')}\n"
        f"Дата/время: {data.get('datetime')}\n"
        f"Комментарий: {data.get('comment')}"
    )

    await state.set_state(OrderForm.waiting_for_confirmation)
    await message.answer(text, reply_markup=confirm_keyboard)


# ✅ ИСПРАВЛЕНИЕ: убираем клавиатуру после нажатия кнопок
@dp.message(OrderForm.waiting_for_confirmation)
async def process_confirmation(message: Message, state: FSMContext, bot: Bot):
    answer = (message.text or "").strip()

    if answer == "Подтвердить":
        data = await state.get_data()

        await state.clear()
        await message.answer(
            "Заявка подтверждена. Спасибо!\n"
            "Мы свяжемся с вами в ближайшее время.",
            reply_markup=ReplyKeyboardRemove(),  # ✅ кнопки пропадают
        )

        text_to_admin = (
            "Новая заявка:\n"
            f"Услуга: {data.get('service')}\n"
            f"Имя: {data.get('name')}\n"
            f"Телефон: {data.get('phone')}\n"
            f"Дата/время: {data.get('datetime')}\n"
            f"Комментарий: {data.get('comment')}"
        )

        await notify_admin(bot, text_to_admin)
        save_order_to_excel(data)
        return

    if answer == "Изменить":
        await state.clear()
        await message.answer(
            "Ок, заполним заявку заново. Выберите услугу:",
            reply_markup=services_keyboard,  # ✅ заменяем клавиатуру на новую
        )
        return

    if answer == "Отменить":
        await state.clear()
        await message.answer(
            "Заявка отменена. Чтобы начать заново, нажмите /start.",
            reply_markup=services_keyboard,  # ✅ заменяем клавиатуру на новую
        )
        return

    await message.answer(
        "Пожалуйста, выберите одну из кнопок:\n"
        "Подтвердить / Изменить / Отменить",
        reply_markup=confirm_keyboard,
    )


# -------------------- ADMIN (ОБНОВЛЁННЫЙ ОДИН РАЗ) --------------------

@dp.message(Command("admin"))
async def admin_panel(message: Message):
    if not message.from_user:
        return
    if not is_admin(message.from_user.id):
        await message.answer("Нет доступа.")
        return

    count = get_orders_count()
    await message.answer(
        f"Админ-панель\n\nКоличество заявок: {count}",
        reply_markup=ADMIN_KEYBOARD,
    )


@dp.message(lambda m: m.text == "Количество заявок")
async def admin_count(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        await message.answer("Нет доступа.")
        return
    await message.answer(f"Количество заявок: {get_orders_count()}")


@dp.message(lambda m: m.text == "Последние заявки")
async def admin_last_orders(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        await message.answer("Нет доступа.")
        return

    orders = load_orders_from_excel(limit=10)
    if not orders:
        await message.answer("Заявок пока нет.")
        return

    lines = ["Последние заявки:"]
    for i, row in enumerate(orders, start=1):
        ts = row[0] if len(row) > 0 else ""
        service = row[1] if len(row) > 1 else ""
        name = row[2] if len(row) > 2 else ""
        phone = row[3] if len(row) > 3 else ""
        dt = row[4] if len(row) > 4 else ""
        comment = row[5] if len(row) > 5 else ""

        lines.append(
            f"{i}) {service}\n"
            f"   {name} | {phone}\n"
            f"   {dt}\n"
            f"   Комм: {comment}\n"
            f"   ({ts})"
        )

    await message.answer("\n".join(lines))


@dp.message(lambda m: m.text == "Выгрузить Excel")
async def admin_export_excel(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        await message.answer("Нет доступа.")
        return

    if not os.path.exists(EXCEL_PATH):
        await message.answer("Excel-файл не найден. Сначала создайте хотя бы одну заявку.")
        return

    await message.answer_document(
        FSInputFile(EXCEL_PATH),
        caption="Выгрузка заявок (Excel)"
    )


@dp.message()
async def fallback(message: Message):
    await message.answer(
        "Я пока понимаю только выбор услуги. Нажмите одну из кнопок ниже.",
        reply_markup=services_keyboard,
    )


async def precheck(bot: Bot):
    info = await bot.get_me()
    logging.info(f"Telegram reachable. Bot username: @{info.username}")


async def set_bot_commands(bot: Bot):
    commands = [
        BotCommand(command="start", description="Открыть меню"),
        BotCommand(command="admin", description="Админ-панель"),
    ]
    await bot.set_my_commands(commands)


async def main():
    PROXY_URL = os.getenv("PROXY_URL", "").strip()

    timeout = ClientTimeout(total=120, connect=60, sock_connect=60, sock_read=60)
    connector = TCPConnector(
        limit=0,
        limit_per_host=0,
        ttl_dns_cache=300,
        enable_cleanup_closed=True,
    )

    bot_kwargs = dict(token=BOT_TOKEN, timeout=timeout, connector=connector)
    if PROXY_URL:
        bot_kwargs["proxy"] = PROXY_URL

    bot = Bot(**bot_kwargs)

    while True:
        try:
            logging.info(f"Excel path: {EXCEL_PATH}")
            logging.info("Precheck: calling getMe()...")
            await precheck(bot)
            await set_bot_commands(bot)
            logging.info("Starting polling...")
            await dp.start_polling(bot)
        except Exception as e:
            logging.exception(f"Polling loop error: {type(e).__name__}: {e}")
            await asyncio.sleep(5)


if __name__ == "__main__":
    asyncio.run(main())
